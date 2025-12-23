from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from django.conf import settings
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST

from .models import Analysis, AnalysisPage, RegionSelection
from .services.cropper import crop_region_from_page
from .services.ai_extractor import extract_design_metadata, extract_bom_materials
from .services.masterfile_builder import (
    append_equipment_to_masterfile,
    parse_filename,
)
from .services.ppt_builder import sync_all_slides_from_masterfile

from core_app.decorators import rbi_login_required
import jwt


STEP_LABEL = {
    "design_data": "Select Design Data Region",
    "bom": "Select Bill Of Material Region",
    "slide_image": "Select Slide Image Region",
}
def _user_key(analysis: Analysis) -> str:
    
    if analysis.created_by:
        return analysis.created_by.staff_id or analysis.created_by.external_id or "anon"
    return "anon"


@rbi_login_required
def upload_analysis(request):
    if request.method == "POST":
        pdf_file = request.FILES.get("pdf_file")
        if not pdf_file:
            messages.error(request, "Please select a PDF file.")
            return redirect("analysis_app:upload")

        
        ext_user = getattr(request, "external_user", None)
        if not ext_user:
            messages.error(request, "Session expired. Please log in again.")
            return redirect("login")

        analysis = Analysis.objects.create(
            created_by=ext_user,
            file=pdf_file,
            original_filename=pdf_file.name,
            status="awaiting_regions",
        )

        pdf_path = analysis.file.path

        from pdf2image import convert_from_path
        POPPLER_PATH = r"C:\Users\nazrisaidon\Desktop\poppler-25.11.0\Library\bin"

        try:
            pages = convert_from_path(pdf_path, dpi=200, poppler_path=POPPLER_PATH)

            for idx, page_img in enumerate(pages, start=1):
                out_dir = Path(settings.MEDIA_ROOT) / "analysis/pages/"
                out_dir.mkdir(parents=True, exist_ok=True)

                filename = f"analysis_{analysis.id}_p{idx}.png"
                out_path = out_dir / filename
                page_img.save(out_path, "PNG")

                rel_path = str(out_path.relative_to(settings.MEDIA_ROOT))

                AnalysisPage.objects.create(
                    analysis=analysis,
                    page_number=idx,
                    image=rel_path,
                )

        except Exception as e:
            analysis.delete()
            print(f"Error converting PDF: {e}")
            messages.error(request, f"Error processing PDF: {e}")
            return redirect("analysis_app:upload")

        return redirect(
            "analysis_app:select_region",
            analysis_id=analysis.id,
            step_type="design_data",
            page_number=1,
        )

    return render(request, "uploading.html")


@rbi_login_required
def select_region(request, analysis_id, step_type, page_number):
 
    if step_type not in ("design_data", "bom", "slide_image"):
        messages.error(request, "Invalid step type.")
        return redirect("analysis_app:history")

    analysis = get_object_or_404(Analysis, pk=analysis_id)
    page = get_object_or_404(AnalysisPage, analysis=analysis, page_number=page_number)


    original_name = analysis.original_filename or ""
    is_h004 = "H-004" in original_name


    all_pages = list(analysis.pages.all())
    page_numbers = [p.page_number for p in all_pages]
    idx = page_numbers.index(page_number)

    prev_page_url = None
    next_page_url = None

    if idx > 0:
        prev_page_url = reverse(
            "analysis_app:select_region",
            kwargs={
                "analysis_id": analysis.id,
                "step_type": step_type,
                "page_number": page_numbers[idx - 1],
            },
        )

    if idx < len(page_numbers) - 1:
        next_page_url = reverse(
            "analysis_app:select_region",
            kwargs={
                "analysis_id": analysis.id,
                "step_type": step_type,
                "page_number": page_numbers[idx + 1],
            },
        )


    if request.method == "POST":

        action = request.POST.get("action", "next")

        x1 = request.POST.get("x1")
        y1 = request.POST.get("y1")
        x2 = request.POST.get("x2")
        y2 = request.POST.get("y2")

        if not all([x1, y1, x2, y2]):
            messages.error(request, "Please draw a region before confirming.")
        else:
            try:
                x1 = float(x1)
                y1 = float(y1)
                x2 = float(x2)
                y2 = float(y2)
            except ValueError:
                messages.error(request, "Invalid coordinates.")
            else:
                if not (
                    0 <= x1 <= 1
                    and 0 <= x2 <= 1
                    and 0 <= y1 <= 1
                    and 0 <= y2 <= 1
                ):
                    messages.error(request, "Coordinates must be between 0 and 1.")
                else:
                  
                    if step_type == "design_data":
                        
                        RegionSelection.objects.update_or_create(
                            analysis=analysis,
                            step_type=step_type,
                            defaults={
                                "page": page,
                                "x1": x1,
                                "y1": y1,
                                "x2": x2,
                                "y2": y2,
                            },
                        )

                    elif step_type == "bom":
                        if is_h004:
                           
                            RegionSelection.objects.create(
                                analysis=analysis,
                                page=page,
                                step_type=step_type,
                                x1=x1,
                                y1=y1,
                                x2=x2,
                                y2=y2,
                            )
                        else:
                           
                            RegionSelection.objects.update_or_create(
                                analysis=analysis,
                                step_type=step_type,
                                defaults={
                                    "page": page,
                                    "x1": x1,
                                    "y1": y1,
                                    "x2": x2,
                                    "y2": y2,
                                },
                            )

                    else:
                        
                        RegionSelection.objects.create(
                            analysis=analysis,
                            page=page,
                            step_type=step_type,
                            x1=x1,
                            y1=y1,
                            x2=x2,
                            y2=y2,
                        )

                    messages.success(request, "Region saved.")

                  
                    if step_type == "design_data":
                        return redirect(
                            "analysis_app:select_region",
                            analysis_id=analysis.id,
                            step_type="bom",
                            page_number=page_number,
                        )

                    elif step_type == "bom":
                     
                        if is_h004 and action == "add_more":
                            return redirect(
                                "analysis_app:select_region",
                                analysis_id=analysis.id,
                                step_type="bom",
                                page_number=page_number,
                            )
                        
                        return redirect(
                            "analysis_app:select_region",
                            analysis_id=analysis.id,
                            step_type="slide_image",
                            page_number=page_number,
                        )

                    else:
                        
                        return redirect(
                            "analysis_app:review_analysis",
                            analysis_id=analysis.id,
                        )

    context = {
        "analysis": analysis,
        "page_number": page.page_number,
        "page_image_url": page.image.url
        if hasattr(page.image, "url")
        else settings.MEDIA_URL + page.image.name,
        "step_type": step_type,
        "step_label": STEP_LABEL.get(step_type, "Select Region"),
        "prev_page_url": prev_page_url,
        "next_page_url": next_page_url,
        "total_pages": len(page_numbers),
        "is_h004": is_h004, 
    }
    return render(request, "select_region.html", context)



@rbi_login_required
def review_analysis(request, analysis_id):
    analysis = get_object_or_404(Analysis, pk=analysis_id)
    regions = analysis.regions.select_related("page").all()

    
    design_data_region = regions.filter(step_type="design_data").first()
    
    bom_regions = list(regions.filter(step_type="bom"))
    
    slide_regions = list(regions.filter(step_type="slide_image"))

   
    design_preview_url = None
    bom_previews: List[Dict[str, Any]] = []
    slide_previews: List[Dict[str, Any]] = []

 
    if design_data_region:
        design_rel = crop_region_from_page(
            page_image_name=design_data_region.page.image.name,
            x1=design_data_region.x1,
            y1=design_data_region.y1,
            x2=design_data_region.x2,
            y2=design_data_region.y2,
        )
        design_preview_url = settings.MEDIA_URL + design_rel

   
    for r in bom_regions[:3]:
        bom_rel = crop_region_from_page(
            page_image_name=r.page.image.name,
            x1=r.x1,
            y1=r.y1,
            x2=r.x2,
            y2=r.y2,
        )
        bom_previews.append(
            {
                "page": r.page.page_number,
                "url": settings.MEDIA_URL + bom_rel,
            }
        )

   
    for r in slide_regions[:3]:
        rel = crop_region_from_page(
            page_image_name=r.page.image.name,
            x1=r.x1,
            y1=r.y1,
            x2=r.x2,
            y2=r.y2,
        )
        slide_previews.append(
            {
                "page": r.page.page_number,
                "url": settings.MEDIA_URL + rel,
            }
        )

    context = {
        "analysis": analysis,
        "design_data_region": design_data_region,
        "bom_regions": bom_regions,
        "slide_regions": slide_regions,
        "design_preview_url": design_preview_url,
        "bom_previews": bom_previews,
        "slide_previews": slide_previews,
    }
    return render(request, "review.html", context)



@rbi_login_required
def analysis_detail(request, analysis_id):
    analysis = get_object_or_404(Analysis, pk=analysis_id)
    regions = analysis.regions.select_related("page").all()

    design_data_region = regions.filter(step_type="design_data").first()
    bom_region = regions.filter(step_type="bom").first()
    slide_regions = regions.filter(step_type="slide_image")


    workbook_url = None
    if analysis.workbook_path:
        workbook_url = settings.MEDIA_URL + analysis.workbook_path

    
    design_rows_preview: List[Dict[str, Any]] = []
    bom_rows_preview: List[Dict[str, Any]] = []

    try:
        if design_data_region:
            _ = crop_region_from_page(
                design_data_region.page.image.name,
                design_data_region.x1,
                design_data_region.y1,
                design_data_region.x2,
                design_data_region.y2,
            )
         
    except Exception as e:
        print("Design data preview error:", e)

    try:
        if bom_region:
            _ = crop_region_from_page(
                bom_region.page.image.name,
                bom_region.x1,
                bom_region.y1,
                bom_region.x2,
                bom_region.y2,
            )
         
    except Exception as e:
        print("BOM preview error:", e)

    context = {
        "analysis": analysis,
        "design_data_region": design_data_region,
        "bom_region": bom_region,
        "slide_regions": slide_regions,
        "workbook_url": workbook_url,
        "design_rows_preview": design_rows_preview,
        "bom_rows_preview": bom_rows_preview,
    }
    return render(request, "detail.html", context)



@rbi_login_required
@require_POST
def generate_analysis(request, analysis_id):
    analysis = get_object_or_404(Analysis, pk=analysis_id)
    regions = analysis.regions.select_related("page").all()

    design_region = regions.filter(step_type="design_data").first()
    bom_regions = list(regions.filter(step_type="bom"))
    slide_regions = list(regions.filter(step_type="slide_image"))

    if not design_region or not bom_regions:
        messages.error(request, "Design Data and BOM regions are required.")
        return redirect("analysis_app:review_analysis", analysis_id=analysis.id)

    analysis.status = "in_progress"
    analysis.save(update_fields=["status"])

    user_key = _user_key(analysis)

    if not analysis.workbook_path:
        analysis.workbook_path = f"analysis/workbooks/{user_key}_IPETRO_Masterfile.xlsx"
        analysis.save(update_fields=["workbook_path"])

    if not analysis.pptx_path:
        analysis.pptx_path = f"analysis/ppt/{user_key}_InspectionPlan.pptx"
        analysis.save(update_fields=["pptx_path"])

    pmt_no, equipment_no = parse_filename(analysis.original_filename)

    design_crop_rel = crop_region_from_page(
        design_region.page.image.name,
        design_region.x1,
        design_region.y1,
        design_region.x2,
        design_region.y2,
    )

    bom_items: List[Dict[str, Any]] = []
    for bom_region in bom_regions:
        bom_crop_rel = crop_region_from_page(
            bom_region.page.image.name,
            bom_region.x1,
            bom_region.y1,
            bom_region.x2,
            bom_region.y2,
        )
        try:
            items = extract_bom_materials(
                bom_crop_rel,
                pmt_no=pmt_no,
                equipment_no=equipment_no,
            ) or []
            bom_items.extend(items)
        except Exception as e:
            print("BOM materials extraction failed for one region:", e)

    design_meta: Dict[str, Any] = {}
    try:
        design_meta = extract_design_metadata(
            design_crop_rel,
            pmt_no=pmt_no,
            equipment_no=equipment_no,
        ) or {}
    except Exception as e:
        print("Design metadata extraction failed:", e)

    slide_image_paths: List[str] = []
    for r in slide_regions:
        crop_rel = crop_region_from_page(
            page_image_name=r.page.image.name,
            x1=r.x1,
            y1=r.y1,
            x2=r.x2,
            y2=r.y2,
        )
        slide_image_paths.append(crop_rel)

    image_map: Dict[Tuple[str, str], str] = {}
    if slide_image_paths:
        image_map[(pmt_no, equipment_no)] = slide_image_paths[0]

    try:
        append_equipment_to_masterfile(
            workbook_rel_path=analysis.workbook_path,
            original_filename=analysis.original_filename,
            design_meta=design_meta,
            bom_items=bom_items,
        )
        sync_all_slides_from_masterfile(
            pptx_rel_path=analysis.pptx_path,
            workbook_rel_path=analysis.workbook_path,
            image_map=image_map or None,
        )
    except Exception as e:
        print("Append to Masterfile failed:", e)

    try:
        sync_all_slides_from_masterfile(
            pptx_rel_path=analysis.pptx_path,
            workbook_rel_path=analysis.workbook_path,
            image_map=image_map or None,
        )
    except Exception as e:
        print("PPT sync failed:", e)

    analysis.status = "awaiting_excel_review"
    analysis.save(update_fields=["status"])

    messages.success(
        request,
        "Draft Excel Masterfile and PowerPoint have been updated from the latest data. "
        "Please review/edit the Excel online.",
    )
    return redirect("analysis_app:analysis_detail", analysis_id=analysis.id)


@rbi_login_required
def edit_masterfile(request, analysis_id):
    analysis = get_object_or_404(Analysis, pk=analysis_id)

    if not analysis.workbook_path:
        messages.error(request, "Masterfile not found for this analysis.")
        return redirect("analysis_app:analysis_detail", analysis_id=analysis.id)

    abs_path = Path(settings.MEDIA_ROOT) / analysis.workbook_path
    if not abs_path.exists():
        messages.error(request, "Workbook file is missing on server.")
        return redirect("analysis_app:analysis_detail", analysis_id=analysis.id)

    wb = openpyxl.load_workbook(abs_path, data_only=True)
    sheet_name = "Masterfile" if "Masterfile" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    DATA_START_ROW = 8  
    MAX_COL = 20        

    rows: List[List[str]] = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row_data: List[str] = []
        is_empty = True
        for col_idx in range(1, MAX_COL + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val not in (None, ""):
                is_empty = False
            row_data.append("" if val is None else str(val))
        if not is_empty:
            rows.append(row_data)

    if not rows:
        
        rows = [["" for _ in range(MAX_COL)] for _ in range(5)]

    table_data_json = json.dumps(rows)

    context = {
        "analysis": analysis,
        "table_data_json": table_data_json,
    }
    return render(request, "edit_masterfile.html", context)


@rbi_login_required
@require_POST
def save_masterfile(request, analysis_id):
    analysis = get_object_or_404(Analysis, pk=analysis_id)

    raw = request.POST.get("table_data")
    if not raw:
        messages.error(request, "No data received from grid.")
        return redirect("analysis_app:edit_masterfile", analysis_id=analysis.id)

    try:
        grid = json.loads(raw)
    except json.JSONDecodeError as e:
        print("SAVE_MASTERFILE JSON ERROR:", e)
        messages.error(request, "Invalid data format from grid.")
        return redirect("analysis_app:edit_masterfile", analysis_id=analysis.id)

    if not analysis.workbook_path:
        messages.error(request, "No workbook attached to this analysis.")
        return redirect("analysis_app:edit_masterfile", analysis_id=analysis.id)

    abs_path = Path(settings.MEDIA_ROOT) / analysis.workbook_path
    if not abs_path.exists():
        messages.error(request, "Workbook file not found on server.")
        return redirect("analysis_app:edit_masterfile", analysis_id=analysis.id)

    wb = load_workbook(abs_path)
    ws = wb["Masterfile"]

    start_row = 8
    start_col = 1

    max_cols = max((len(r) for r in grid), default=0)
    for r in grid:
        while len(r) < max_cols:
            r.append("")

    old_last_row = ws.max_row
    for row_idx in range(start_row, old_last_row + 1):
        for col_idx in range(start_col, start_col + max_cols):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None

    for r_idx, row_data in enumerate(grid, start=start_row):
        for c_idx, value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = value

    wb.save(abs_path)

    try:
        user_key = _user_key(analysis)

        if not analysis.pptx_path:
            analysis.pptx_path = f"analysis/ppt/{user_key}_InspectionPlan.pptx"
            analysis.save(update_fields=["pptx_path"])

        sync_all_slides_from_masterfile(
            pptx_rel_path=analysis.pptx_path,
            workbook_rel_path=analysis.workbook_path,
            image_map=None,
        )
    except Exception as e:
        print("PPT sync failed after save_masterfile:", e)
        messages.warning(request, "✅ Masterfile saved, but PPT sync failed.")
        return redirect("analysis_app:edit_masterfile", analysis_id=analysis.id)

    messages.success(request, "✅ Masterfile has been saved successfully.")
    return redirect("analysis_app:edit_masterfile", analysis_id=analysis.id)


@rbi_login_required
@require_POST
def upload_corrected_masterfile(request, analysis_id):
    analysis = get_object_or_404(Analysis, pk=analysis_id)

    file_obj = request.FILES.get("masterfile")
    if not file_obj:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect("analysis_app:analysis_detail", analysis_id=analysis.id)

    if not file_obj.name.lower().endswith((".xlsx", ".xlsm")):
        messages.error(request, "Invalid file type. Please upload an .xlsx or .xlsm file.")
        return redirect("analysis_app:analysis_detail", analysis_id=analysis.id)

    if not analysis.workbook_path:
        user_key = _user_key(analysis)
        analysis.workbook_path = f"analysis/workbooks/{user_key}_IPETRO_Masterfile.xlsx"
        analysis.save(update_fields=["workbook_path"])

    abs_path = Path(settings.MEDIA_ROOT) / analysis.workbook_path
    abs_path.parent.mkdir(parents=True, exist_ok=True)

    with abs_path.open("wb+") as dest:
        for chunk in file_obj.chunks():
            dest.write(chunk)

    analysis.status = "completed"
    analysis.save(update_fields=["status"])

    messages.success(request, "Corrected Masterfile uploaded successfully.")
    return redirect("analysis_app:analysis_detail", analysis_id=analysis.id)


@rbi_login_required
def analysis_history(request):
    qs = Analysis.objects.all().order_by("-created_at")

    ext_user = getattr(request, "external_user", None)
    if ext_user:
        qs = qs.filter(created_by=ext_user)

    paginator = Paginator(qs, 10)
    page_number = request.GET.get("page")
    analyses = paginator.get_page(page_number)

    return render(request, "history.html", {"analyses": analyses})
