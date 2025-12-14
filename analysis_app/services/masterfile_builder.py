# analysis_app/services/masterfile_builder.py
from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from copy import copy 


from django.conf import settings
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from .material_utils import parse_spec_grade as parse_material



# --- Configuration ----------------------------------------------------------------

MASTERFILE_TEMPLATE_PATH = (
    Path(settings.BASE_DIR)
    / "rbi_templates"
    / "MasterFile _ IPETRO PLANT.xlsx"
)

MASTERFILE_SHEET_NAME = "Masterfile"

# Equipment yang OPERATING dia WAJIB ikut template (H-001 .. H-004)
USE_TEMPLATE_OPERATING = {
    ("MLK PMT 10107", "H-001"),
    ("MLK PMT 10108", "H-002"),
    ("MLK PMT 10109", "H-003"),
    ("MLK PMT 10110", "H-004"),
}

# Column indices (ikut Excel awak)
COL_NO = 1
COL_EQUIPMENT_NO = 2
COL_PMT_NO = 3
COL_DESCRIPTION = 4
COL_PARTS = 5
COL_PHASE = 6
COL_FLUID = 7
COL_TYPE = 8
COL_SPEC = 9
COL_GRADE = 10
COL_INSULATION = 11
COL_DESIGN_TEMP = 12
COL_DESIGN_PRESS = 13
COL_OPER_TEMP = 14
COL_OPER_PRESS = 15

FIRST_DATA_ROW = 8  



def copy_row_style(ws, source_row: int, target_row: int, max_col: int = COL_OPER_PRESS) -> None:
  
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)

        if getattr(src, "has_style", False):
            tgt._style = copy(src._style)

@dataclass
class TemplatePartPattern:
    description: str
    part: str
    phase: Optional[str]
    type_name: Optional[str]
    oper_temp: Optional[float] = None   
    oper_press: Optional[float] = None  



def _ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def parse_filename(original_filename: str) -> Tuple[str, str]:
   
    from pathlib import Path as _P

    stem = _P(original_filename).stem  
    if " - " in stem:
        pmt_part, eq_part = [s.strip() for s in stem.split(" - ", 1)]
    else:
        tokens = stem.split()
        if not tokens:
            return stem, ""
        eq_part = tokens[-1]
        pmt_part = " ".join(tokens[:-1])
    return pmt_part, eq_part

def _norm_pmt(s: str) -> str:
    
    return " ".join((s or "").strip().upper().split())


def _norm_eq(s: str) -> str:
   
    return (s or "").strip().upper().replace(" ", "")


def get_design_rule(pmt_no: str, equipment_no: str) -> Optional[DesignTemplateRule]:
    p_norm = _norm_pmt(pmt_no)
    e_norm = _norm_eq(equipment_no)
    for (p, e), rule in DESIGN_RULES.items():
        if _norm_pmt(p) == p_norm and _norm_eq(e) == e_norm:
            return rule
    return None


def get_bom_rule(pmt_no: str, equipment_no: str) -> Optional[BomTemplateRule]:
    p_norm = _norm_pmt(pmt_no)
    e_norm = _norm_eq(equipment_no)
    for (p, e), rule in BOM_RULES.items():
        if _norm_pmt(p) == p_norm and _norm_eq(e) == e_norm:
            return rule
    return None

def _use_template_operating(pmt_no: str, equipment_no: str) -> bool:

    key_pmt = _norm_pmt(pmt_no)
    key_eq = _norm_eq(equipment_no)
    for p, e in USE_TEMPLATE_OPERATING:
        if _norm_pmt(p) == key_pmt and _norm_eq(e) == key_eq:
            return True
    return False


def load_masterfile_template():
  
    if not MASTERFILE_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Masterfile template not found at {MASTERFILE_TEMPLATE_PATH}. "
            f"Please put 'MasterFile _ IPETRO PLANT.xlsx' there or update MASTERFILE_TEMPLATE_PATH."
        )
    wb = load_workbook(MASTERFILE_TEMPLATE_PATH, data_only=True)
    ws = wb[MASTERFILE_SHEET_NAME]
    return wb, ws


def extract_equipment_pattern(
    ws_template,
    pmt_no: str,
    equipment_no: str,
) -> Tuple[List[TemplatePartPattern], Optional[float], Optional[float]]:
  
    patterns: List[TemplatePartPattern] = []
    max_row = ws_template.max_row

    
    pmt_no_norm = _norm_pmt(pmt_no)
    equipment_no_norm = _norm_eq(equipment_no)

    for row in range(FIRST_DATA_ROW, max_row + 1):
        eq_val = (ws_template.cell(row=row, column=COL_EQUIPMENT_NO).value or "")
        pmt_val = (ws_template.cell(row=row, column=COL_PMT_NO).value or "")

        eq_norm = _norm_eq(eq_val)
        pmt_norm = _norm_pmt(pmt_val)

        if eq_norm == equipment_no_norm and pmt_norm == pmt_no_norm:
            description = ws_template.cell(row=row, column=COL_DESCRIPTION).value or ""
            current_row = row

            while current_row <= max_row:
                eq2 = (ws_template.cell(current_row, COL_EQUIPMENT_NO).value or "").strip()
                pmt2 = (ws_template.cell(current_row, COL_PMT_NO).value or "").strip()
                if current_row != row and (eq2 or pmt2):
                  
                    break

                
                raw_part = ws_template.cell(current_row, COL_PARTS).value
                part = ""
                if raw_part is not None:
                    part = str(raw_part).strip()
                if not part:
                    break

                
                phase_raw = ws_template.cell(current_row, COL_PHASE).value
                type_raw = ws_template.cell(current_row, COL_TYPE).value
                phase = str(phase_raw).strip() if phase_raw not in (None, "") else None
                type_name = str(type_raw).strip() if type_raw not in (None, "") else None

                
                t_val = ws_template.cell(current_row, COL_OPER_TEMP).value
                p_val = ws_template.cell(current_row, COL_OPER_PRESS).value
                try:
                    row_oper_temp = float(t_val) if t_val not in ("", None) else None
                except (TypeError, ValueError):
                    row_oper_temp = None
                try:
                    row_oper_press = float(p_val) if p_val not in ("", None) else None
                except (TypeError, ValueError):
                    row_oper_press = None

                patterns.append(
                    TemplatePartPattern(
                        description=str(description),
                        part=part,
                        phase=phase,
                        type_name=type_name,
                        oper_temp=row_oper_temp,
                        oper_press=row_oper_press,
                    )
                )
                current_row += 1
            break

   
    tmpl_oper_temp = next((p.oper_temp for p in patterns if p.oper_temp is not None), None)
    tmpl_oper_press = next((p.oper_press for p in patterns if p.oper_press is not None), None)

    print("DEBUG patterns count:", len(patterns), "for", pmt_no, "/", equipment_no)
    print("DEBUG tmpl_oper_temp / tmpl_oper_press from template:", tmpl_oper_temp, tmpl_oper_press)
    for p in patterns:
        print(
            "  pattern:", repr(p.part),
            "| oper_temp:", p.oper_temp,
            "| oper_press:", p.oper_press,
        )

    return patterns, tmpl_oper_temp, tmpl_oper_press



def get_next_no(ws) -> int:
   
    max_no = 0
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        val = ws.cell(row=row, column=COL_NO).value
        try:
            if val is not None:
                num = int(val)
                if num > max_no:
                    max_no = num
        except (TypeError, ValueError):
            continue
    return max_no + 1 if max_no > 0 else 1


def find_first_empty_data_row(ws) -> int:
  
    last_row = ws.max_row

    for row in range(FIRST_DATA_ROW, last_row + 1):
        is_empty = True
        for c in range(1, COL_OPER_PRESS + 1):
            cell = ws.cell(row=row, column=c)
            if cell.value not in (None, ""):
                is_empty = False
                break
        if is_empty:
            return row

    return last_row + 1


def infer_side_from_part(part_label: str) -> str:
    
    p = (part_label or "").lower()
    if any(k in p for k in ("tube", "bundle", "channel", "header")):
        return "tube"
    return "shell"


def _normalise_token(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def find_best_material_for_part(
    bom_items: List[Dict[str, Any]],
    part_label: str,
) -> Optional[Dict[str, Any]]:
    
    if not bom_items:
        return None

    norm_target = _normalise_token(part_label)
    side_target = infer_side_from_part(part_label)

    
    for item in bom_items:
        label = item.get("part_label") or ""
        if _normalise_token(label) == norm_target:
            return item

   
    candidates: List[Dict[str, Any]] = []
    for item in bom_items:
        label = item.get("part_label") or ""
        norm_label = _normalise_token(label)
        if norm_label and (norm_label in norm_target or norm_target in norm_label):
            candidates.append(item)
    if candidates:
        return candidates[0]

  
    for item in bom_items:
        side = (item.get("side") or "").lower()
        if side and side == side_target:
            return item

    
    return bom_items[0]


def _normalise_insulation(raw_insulation: Optional[str]) -> Optional[str]:
  
    if raw_insulation is None:
        return "NO"

    text = str(raw_insulation).strip()
    if not text:
        return "NO"

    upper = text.upper()

    if upper in {"0", "-", "NIL", "NONE", "NO INSULATION"}:
        return "NO"

    return "YES"


def get_or_create_masterfile_workbook(workbook_rel_path: str) -> Path:
    
  
    media_root = Path(settings.MEDIA_ROOT)
    abs_path = media_root / workbook_rel_path

   
    if abs_path.exists():
        wb_out = load_workbook(abs_path)
        ws_out = wb_out[MASTERFILE_SHEET_NAME]

       
        if ws_out.max_row >= FIRST_DATA_ROW:
           
            template_row = FIRST_DATA_ROW
            for r in range(FIRST_DATA_ROW, ws_out.max_row + 1):
                copy_row_style(ws_out, template_row, r, COL_OPER_PRESS)

         
            base_height = ws_out.row_dimensions[FIRST_DATA_ROW].height or 15
            for r in range(FIRST_DATA_ROW, ws_out.max_row + 1):
                ws_out.row_dimensions[r].height = base_height

            wb_out.save(abs_path)

        return abs_path

   
    if not MASTERFILE_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Masterfile template not found at {MASTERFILE_TEMPLATE_PATH}. "
            "Please put the original template there or update MASTERFILE_TEMPLATE_PATH."
        )

    _ensure_parent_dir(abs_path)
    shutil.copy2(MASTERFILE_TEMPLATE_PATH, abs_path)

    wb_out = load_workbook(abs_path)
    ws_out = wb_out[MASTERFILE_SHEET_NAME]

    
    ranges_to_remove = [
        rng for rng in ws_out.merged_cells.ranges if rng.min_row >= FIRST_DATA_ROW
    ]
    for rng in list(ranges_to_remove):
        ws_out.unmerge_cells(str(rng))

 
    for row in range(FIRST_DATA_ROW, ws_out.max_row + 1):
        for col in range(1, COL_OPER_PRESS + 1):
            ws_out.cell(row=row, column=col).value = None

    
    if ws_out.max_row >= FIRST_DATA_ROW:
        template_row = FIRST_DATA_ROW
        for r in range(FIRST_DATA_ROW, ws_out.max_row + 1):
            copy_row_style(ws_out, template_row, r, COL_OPER_PRESS)

       
        base_height = ws_out.row_dimensions[FIRST_DATA_ROW].height or 15
        for r in range(FIRST_DATA_ROW, ws_out.max_row + 1):
            ws_out.row_dimensions[r].height = base_height

    wb_out.save(abs_path)
    return abs_path


def append_equipment_to_masterfile(
    workbook_rel_path: str,
    original_filename: str,
    design_meta: Dict[str, Any],
    bom_items: List[Dict[str, Any]],
) -> None:

    pmt_no, equipment_no = parse_filename(original_filename)
    print("DEBUG append_equipment_to_masterfile for:", pmt_no, "/", equipment_no)

   
    _tmpl_wb, tmpl_ws = load_masterfile_template()
    patterns, tmpl_oper_temp, tmpl_oper_press = extract_equipment_pattern(
        tmpl_ws,
        pmt_no,
        equipment_no,
    )

    if not patterns:
        print(f"[Masterfile] No template pattern found for {pmt_no} / {equipment_no}")
        return

   
    abs_path = get_or_create_masterfile_workbook(workbook_rel_path)
    wb_out = load_workbook(abs_path)
    ws_out = wb_out[MASTERFILE_SHEET_NAME]

    next_no = get_next_no(ws_out)
    current_row = find_first_empty_data_row(ws_out)
    print("DEBUG first empty row:", current_row)

   
    fluids = (design_meta.get("fluids") or {})
    design = (design_meta.get("design") or {})
    operating = (design_meta.get("operating") or {})
    insulation_norm = _normalise_insulation(design_meta.get("insulation"))

    def get_design_for_side(side: str) -> Tuple[Optional[float], Optional[float]]:
        side_block = (design.get(side) or {})
        return side_block.get("temp_c"), side_block.get("pressure_mpa")

    def get_oper_for_side(side: str) -> Tuple[Optional[float], Optional[float]]:
        side_block = (operating.get(side) or {})
        return side_block.get("temp_c"), side_block.get("pressure_mpa")

    
    is_first_row_for_equipment = True
    first_row_for_equipment: Optional[int] = None

    use_template_oper = _use_template_operating(pmt_no, equipment_no)

    for pattern in patterns:
        part_label = pattern.part
        side = infer_side_from_part(part_label)

      
        material_item = find_best_material_for_part(bom_items, part_label) if bom_items else None
        material_raw = material_item.get("material_raw") if material_item else ""
        spec, grade = parse_material(material_raw or "")

        
        if side == "shell":
            fluid_val = (
                fluids.get("shell")
                or fluids.get("shell side")
                or fluids.get("shell_side")
            )
        else:
            fluid_val = (
                fluids.get("tube")
                or fluids.get("header")
                or fluids.get("tube side")
                or fluids.get("tube_side")
            )

        
        des_temp, des_press = get_design_for_side(side)
        if not des_temp and not des_press:
            des_temp, des_press = get_design_for_side("shell")

        
        if use_template_oper:
            
            op_temp = pattern.oper_temp
            op_press = pattern.oper_press

            
            if op_temp is None and op_press is None:
                op_temp, op_press = get_oper_for_side(side)
                if not op_temp and not op_press:
                    op_temp, op_press = get_oper_for_side("shell")
        else:
           
            op_temp, op_press = get_oper_for_side(side)
            if not op_temp and not op_press:
                op_temp, op_press = get_oper_for_side("shell")

        
        r = current_row

        if is_first_row_for_equipment:
            first_row_for_equipment = r

            ws_out.cell(row=r, column=COL_NO).value = next_no
            ws_out.cell(row=r, column=COL_EQUIPMENT_NO).value = equipment_no
            ws_out.cell(row=r, column=COL_PMT_NO).value = pmt_no
            ws_out.cell(row=r, column=COL_DESCRIPTION).value = pattern.description

            is_first_row_for_equipment = False

        ws_out.cell(row=r, column=COL_PARTS).value = part_label
        ws_out.cell(row=r, column=COL_PHASE).value = pattern.phase
        ws_out.cell(row=r, column=COL_FLUID).value = fluid_val or None
        ws_out.cell(row=r, column=COL_TYPE).value = pattern.type_name
        ws_out.cell(row=r, column=COL_SPEC).value = spec or None
        ws_out.cell(row=r, column=COL_GRADE).value = grade or None
        ws_out.cell(row=r, column=COL_INSULATION).value = insulation_norm
        ws_out.cell(row=r, column=COL_DESIGN_TEMP).value = des_temp
        ws_out.cell(row=r, column=COL_DESIGN_PRESS).value = des_press
        ws_out.cell(row=r, column=COL_OPER_TEMP).value = op_temp
        ws_out.cell(row=r, column=COL_OPER_PRESS).value = op_press

        current_row += 1

    if first_row_for_equipment is not None and current_row - first_row_for_equipment > 1:
        last_row_for_equipment = current_row - 1
        for col in (COL_NO, COL_EQUIPMENT_NO, COL_PMT_NO, COL_DESCRIPTION):
            ws_out.merge_cells(
                start_row=first_row_for_equipment,
                start_column=col,
                end_row=last_row_for_equipment,
                end_column=col,
            )

        
        base_height = ws_out.row_dimensions[FIRST_DATA_ROW].height or 15
        for r in range(first_row_for_equipment, last_row_for_equipment + 1):
            ws_out.row_dimensions[r].height = base_height


        desc_cell = ws_out.cell(row=first_row_for_equipment, column=COL_DESCRIPTION)
        desc_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=False,   
        )
    wb_out.save(abs_path)
