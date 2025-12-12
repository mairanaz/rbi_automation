# analysis_app/services/excel_builder.py

from pathlib import Path
from typing import List, Dict
from django.conf import settings
from openpyxl import Workbook, load_workbook


def get_or_create_workbook(workbook_rel_path: str) -> str:
    abs_path = Path(settings.MEDIA_ROOT) / workbook_rel_path
    abs_path.parent.mkdir(parents=True, exist_ok=True)

    if not abs_path.exists():
        wb = Workbook()
        wb.save(abs_path)
    return str(abs_path)


def append_rows_to_sheet(workbook_abs_path: str, sheet_name: str, rows: List[Dict]):
    if not rows:
        return

    abs_path = Path(workbook_abs_path)
    wb = load_workbook(abs_path)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    
    if ws.max_row == 1 and ws["A1"].value is None:
        headers = list(rows[0].keys())
        ws.append(headers)

  
    headers = [cell.value for cell in ws[1]]
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    wb.save(abs_path)
